import datetime

from django.core.exceptions import ValidationError
from openpyxl import load_workbook


class ParserToDatabase:
    """
        Класс для парсинга файлов "Я - как проект"
    """

    _sheetlists_rules = [
            ['ОБЩИЙ ПЛАН', 'Образование', 'Дело',
             'Организация и коллектив', 'Репутация',
             'Здоровье',
             'Семья и окружение', 'Файл'],
            ['ОБЩИЙ ПЛАН', 'Образование по УП', 'Программа СРС (на сем.)',
             'План ПрофРазвития (1 год)', 'Забота о себе']
        ]

    def __init__(self, document):
        self._document = document
        self._path = document.get_file_path()
        self.wb = load_workbook(self._path, read_only=True)
        self.sheetlist = self.wb.sheetnames
        self.event = document.event
        self.rule = {
            'day': self.event.day,
            'month': self.event.month,
            'year': self.event.year
        }

    def check_sheet(self):
        """
        Валидация файла по имеющимся листам
        """

        sheetlist = list(set(self.sheetlist))
        sheetlist.sort(key=lambda x: x)
        sheet_rules = self._sheetlists_rules
        [rule.sort(key=lambda x: x) for rule in sheet_rules]

        if sheetlist not in sheet_rules:
            self.wb.close()
            try:
                self._document.delete()
            except:
                pass

            raise ValidationError('Нереверный макет файла!')
        return self.sheetlist

    def str_to_datetime(self, date_time_str: str):
        try:
            return datetime.datetime.strptime(date_time_str, '%d-%m-%Y') - datetime.datetime(self.rule.get('year', 0),
                                                                                             self.rule.get('month', 0),
                                                                                             self.rule.get('month', 0),
                                                                                             0, 0)
        except:
            return datetime.timedelta(days=0)

    def get_values_by_sheet(self):
        """
        Подсчет очков для каждой страницы
        """

        checklist = {}
        result = {}
        page_variant = -1 if 'Файлы' or 'файлы' in self.sheetlist else len(self.sheetlist)
        print(page_variant, self.sheetlist)
        for page in self.check_sheet()[1:page_variant]:
            for v, _, _, _, _, _, c, i in self.wb[page]['H37':'O56']:

                if not checklist.get(str(page)):
                    checklist[str(page)] = {
                        'Задача': {v.value},
                        'Дедлайн': [c.value - datetime.datetime(self.rule.get('year', 0),
                                                                self.rule.get('month', 0),
                                                                self.rule.get('month', 0),
                                                                0, 0)
                                    if isinstance(c.value, datetime.datetime)
                                    else self.str_to_datetime(c.value)],
                        'Статус': [i.value]
                    }
                else:
                    checklist[str(page)]['Задача'].add(v.value)
                    checklist[str(page)]['Дедлайн'].append(
                        c.value - datetime.datetime(self.rule.get('year', 0),
                                                    self.rule.get('month', 0),
                                                    self.rule.get('month', 0),
                                                    0, 0)
                        if isinstance(c.value, datetime.datetime) else self.str_to_datetime(c.value))
                    checklist[str(page)]['Статус'].append(i.value)

            checklist[page]['Задача'].discard("' '")
            checklist[page]['Задача'].discard(None)

            result[page] = {
                'total_deadline': sum(checklist.get(page)['Дедлайн'], datetime.timedelta(0, 0)),
                'statuses': {i: checklist.get(page)['Статус'].count(i) for i in set(checklist.get(page)['Статус'])},
                'range_task': len(checklist.get(page)['Задача'])
            }
        self.wb.close()

        return result

    def get_sum_of_status(self, values: dict):
        """
        Подсчет итоговых баллов по статусу, в соответствии с правилами
        """

        points = {
            None: 0,
            'Выполнено': 2,
            'В процессе': 1,
            'Не начато': 0,
        }

        counter = 0

        for key in values.keys():
            for k, v in values.get(key)['statuses'].items():
                counter += points.get(k, 0) * v
        return counter

    def get_total_values(self):
        """
        Возвращает суммарные очки за все разделы
        """
        sheet_values_list = self.get_values_by_sheet()
        res = {'deadline': 0, 'status': self.get_sum_of_status(sheet_values_list), 'task': 0}

        for i in sheet_values_list.keys():
            res['deadline'] += sheet_values_list[i]['total_deadline'].days
            res['task'] += sheet_values_list[i]['range_task']

        return res
